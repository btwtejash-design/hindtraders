import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any

def generate_hind_quotation_excel(data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hind Traders Quotation"

    common = data.get("common", {})
    hind_data = data.get("hind_traders", {})

    ref_no = common.get("ref_no", "F/DPS/MMC(D)/27")
    ref_date = common.get("ref_date", "05/2026")
    consignee = common.get("consignee_address", "AWM (WHEEL)\nEASTERN RLY, JAMALPUR")
    org_ref = hind_data.get("quotation_ref", "HT/BQ/26-27")
    org_date = hind_data.get("quotation_date", "03/08/2026")

    items = common.get("items", [])
    rates = hind_data.get("rates", {})

    # Title / Header
    ws.merge_cells("A1:D1")
    ws["A1"] = "HIND TRADERS"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Engineering & Machineries | GSTIN: 10DFIPK1994B1ZS | Mob: 7903235877"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:D3")
    ws["A3"] = "BUDGETARY QUOTATION"
    ws["A3"].font = Font(name="Arial", size=12, bold=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws["A5"] = f"Quotation Ref: {org_ref}"
    ws["D5"] = f"Date: {org_date}"
    ws["A5"].font = Font(bold=True)
    ws["D5"].font = Font(bold=True)
    ws["D5"].alignment = Alignment(horizontal="right")

    ws["A6"] = f"TO:\n{consignee}"
    ws["A6"].font = Font(bold=True)

    ws["A8"] = f"Ref No: {ref_no}"
    ws["D8"] = f"Date: {ref_date}"
    ws["A8"].font = Font(bold=True)
    ws["D8"].font = Font(bold=True)
    ws["D8"].alignment = Alignment(horizontal="right")

    # Table Header
    headers = ["Sl. No", "Description", "Qty", "Rate (₹)"]
    ws.append([]) # Row 9 blank
    header_row = 10
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    row_num = 11
    for idx, item in enumerate(items, 1):
        sr = item.get("sr_no", idx)
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        unit = item.get("unit", "mtr")
        rate_val = rates.get(str(sr)) or rates.get(sr) or 0

        ws.cell(row=row_num, column=1, value=sr).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=3, value=f"{qty} {unit}").alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=float(rate_val) if rate_val else 0).alignment = Alignment(horizontal="right")

        for c in range(1, 5):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).font = Font(name="Arial", size=10, bold=True)
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="Terms & Conditions:\n1. GST@18% Extra\n2. For Destination\n3. Delivery within 30 days\n4. Material guaranteed as per IRS terms & conditions").font = Font(size=9, italic=True)
    ws.cell(row=row_num, column=4, value="For HIND TRADERS\n\nProprietor").font = Font(bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="right")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path

def generate_yasha_quotation_excel(data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yasha Enterprises Quotation"

    common = data.get("common", {})
    yasha_data = data.get("yasha_enterprises", {})

    ref_no = common.get("ref_no", "F/DPS/MMC(D)/27")
    ref_date = common.get("ref_date", "08/2026")
    consignee = common.get("consignee_address", "Thc, Sr. DME,\nEASTERN RAILWAY\nMALDA")
    org_ref = yasha_data.get("quotation_ref", "YE/BQ/26-27")
    org_date = yasha_data.get("quotation_date", "04/08/2026")

    items = common.get("items", [])
    rates = yasha_data.get("rates", {})

    ws.merge_cells("A1:D1")
    ws["A1"] = "YASHA ENTERPRISES"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="880E4F")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Manufacturer of Wagon Components for Indian Railways | GSTIN: 10AOCPS5660Q1ZK"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = f"Ref :- {org_ref}"
    ws["D4"] = f"Date;- {org_date}"
    ws["A4"].font = Font(bold=True)
    ws["D4"].font = Font(bold=True)
    ws["D4"].alignment = Alignment(horizontal="right")

    ws["A5"] = f"TO:\n{consignee}"
    ws["A5"].font = Font(bold=True)

    ws["A7"] = f"Ref:- {ref_no}"
    ws["D7"] = f"Dated:- {ref_date}"
    ws["A7"].font = Font(bold=True)
    ws["D7"].font = Font(bold=True)
    ws["D7"].alignment = Alignment(horizontal="right")

    ws["A8"] = "Dear Sir, We are submitting our best competitive price for this subject material:"
    ws["A8"].font = Font(italic=True)

    headers = ["Sl. No", "Description", "QTY", "RATE (₹)"]
    header_row = 10
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

    row_num = 11
    for idx, item in enumerate(items, 1):
        sr = item.get("sr_no", idx)
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        unit = item.get("unit", "Mtr")
        rate_val = rates.get(str(sr)) or rates.get(sr) or 0

        ws.cell(row=row_num, column=1, value=sr).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=3, value=f"{qty} {unit}").alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=f"{rate_val}/-").alignment = Alignment(horizontal="right")
        row_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path

def generate_madhu_quotation_excel(data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Madhu Enterprises Quotation"

    common = data.get("common", {})
    madhu_data = data.get("madhu_enterprises", {})

    ref_no = common.get("ref_no", "F/DPS/MMC(D)/27")
    ref_date = common.get("ref_date", "28/04/2026")
    consignee = common.get("consignee_address", "The, AWM (DIESEL)\nEASTERN RLY, JAMALPUR")
    org_ref = madhu_data.get("quotation_ref", "ME/12/26-27")
    org_date = madhu_data.get("quotation_date", "06/05/2026")

    items = common.get("items", [])
    rates = madhu_data.get("rates", {})

    ws.merge_cells("A1:D1")
    ws["A1"] = "MADHU ENTERPRISES"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="2E7D32")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "MADHOPUR, NAYA TOLA, MUNGER - 811202 | GSTIN: 10AXZPS6246GIZ0 | Mob: 9471903155"
    ws["A2"].font = Font(name="Arial", size=9, color="2E7D32", italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = f"Ref. No :- {org_ref}"
    ws["D4"] = f"Date . {org_date}"
    ws["A4"].font = Font(bold=True, color="2E7D32")
    ws["D4"].font = Font(bold=True, color="2E7D32")
    ws["D4"].alignment = Alignment(horizontal="right")

    ws["A5"] = f"TO:\n{consignee}"
    ws["A5"].font = Font(bold=True, color="2E7D32")

    ws["A7"] = f"Ref. No :- {ref_no}"
    ws["D7"] = f"Date:- {ref_date}"
    ws["A7"].font = Font(bold=True, color="2E7D32")
    ws["D7"].font = Font(bold=True, color="2E7D32")
    ws["D7"].alignment = Alignment(horizontal="right")

    headers = ["Sl. No", "Description", "Qty", "Rate (₹)"]
    header_row = 9
    green_border = Border(
        left=Side(style='thin', color='2E7D32'),
        right=Side(style='thin', color='2E7D32'),
        top=Side(style='thin', color='2E7D32'),
        bottom=Side(style='thin', color='2E7D32')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")
        cell.border = green_border

    row_num = 10
    for idx, item in enumerate(items, 1):
        sr = item.get("sr_no", idx)
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        unit = item.get("unit", "Nos")
        rate_val = rates.get(str(sr)) or rates.get(sr) or 0

        ws.cell(row=row_num, column=1, value=sr).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=3, value=f"{qty} {unit}").alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=float(rate_val) if rate_val else 0).alignment = Alignment(horizontal="right")

        for c in range(1, 5):
            ws.cell(row=row_num, column=c).border = green_border
            ws.cell(row=row_num, column=c).font = Font(name="Arial", size=10, bold=True, color="2E7D32")
        row_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path


def generate_lovely_quotation_excel(data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lovely Supplier Quotation"

    common = data.get("common", {})
    lovely_data = data.get("lovely_supplier", {})

    ref_no = common.get("ref_no", "F/DPS/MMC(D)/27")
    ref_date = common.get("ref_date", "28/04/2026")
    consignee = common.get("consignee_address", "The, AWM (DIESEL)\nEASTERN RLY, JAMALPUR")
    org_ref = lovely_data.get("quotation_ref", "LV/23/26-27")
    org_date = lovely_data.get("quotation_date", "29/04/2026")

    items = common.get("items", [])
    rates = lovely_data.get("rates", {})

    ws.merge_cells("A1:D1")
    ws["A1"] = "LOVELY GENERAL ORDER SUPPLIER"
    ws["A1"].font = Font(name="Times New Roman", size=16, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "GSTIN: 10EOBPK6340Q1ZU | Vender Code: 57722 | Mob: 9852949143"
    ws["A2"].font = Font(name="Times New Roman", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:D3")
    ws["A3"] = "Budgetary Quotation"
    ws["A3"].font = Font(name="Times New Roman", size=12, bold=True, underline="single")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws["A5"] = f"Ref:- {org_ref}"
    ws["D5"] = f"Date:- {org_date}"
    ws["A5"].font = Font(name="Times New Roman", bold=True, italic=True)
    ws["D5"].font = Font(name="Times New Roman", bold=True, italic=True)
    ws["D5"].alignment = Alignment(horizontal="right")

    ws["A6"] = f"TO:\n{consignee}"
    ws["A6"].font = Font(name="Times New Roman", bold=True)

    ws["A8"] = f"Ref:- {ref_no}"
    ws["D8"] = f"Date:- {ref_date}"
    ws["A8"].font = Font(name="Times New Roman", bold=True, italic=True)
    ws["D8"].font = Font(name="Times New Roman", bold=True, italic=True)
    ws["D8"].alignment = Alignment(horizontal="right")

    headers = ["[Sl. No", "Description", "Qty", "Rate (₹)"]
    header_row = 10
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

    row_num = 11
    for idx, item in enumerate(items, 1):
        sr = item.get("sr_no", idx)
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        unit = item.get("unit", "Nos")
        rate_val = rates.get(str(sr)) or rates.get(sr) or 0

        ws.cell(row=row_num, column=1, value=sr).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=3, value=f"{qty} {unit}").alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=float(rate_val) if rate_val else 0).alignment = Alignment(horizontal="right")

        for c in range(1, 5):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).font = Font(name="Times New Roman", size=10, bold=True)
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="Terms & Condition:\n1. GST@18% Extra\n2. For Destination\n3. Delivery within 30 days").font = Font(name="Times New Roman", size=9, italic=True)
    ws.cell(row=row_num, column=4, value="LOVELY GENERAL ORDER SUPPLIER\nManish Kumar (Proprietor)").font = Font(name="Times New Roman", bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="right")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path


def generate_raju_quotation_excel(data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raju Engg Quotation"

    common = data.get("common", {})
    raju_data = data.get("raju_engineering_works", {})

    ref_no = common.get("ref_no", "Nil")
    ref_date = common.get("ref_date", "Nil")
    consignee = common.get("consignee_address", "Dy, CMT\nEASTERN RLY. JAMALPUR")
    org_ref = raju_data.get("quotation_ref", "REW/BQ/26-27")
    org_date = raju_data.get("quotation_date", "02/06/2026")

    items = common.get("items", [])
    rates = raju_data.get("rates", {})

    ws.merge_cells("A1:D1")
    ws["A1"] = "M/S RAJU ENGINEERING WORKS"
    ws["A1"].font = Font(name="Courier New", size=15, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Railway Contractor | GSTIN: 10JRHPK4490P1Z8 | MSME: UDYAM-BR-22-4015162 | Mob: 8651757734"
    ws["A2"].font = Font(name="Courier New", size=8, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = f"REF.- {org_ref}"
    ws["D4"] = f"DATE:- {org_date}"
    ws["A4"].font = Font(name="Courier New", bold=True)
    ws["D4"].font = Font(name="Courier New", bold=True)
    ws["D4"].alignment = Alignment(horizontal="right")

    ws["A5"] = f"To.\n{consignee}"
    ws["A5"].font = Font(name="Courier New", bold=True)

    ws["A7"] = f"REF. No:- {ref_no}"
    ws["D7"] = f"DATE:- {ref_date}"
    ws["A7"].font = Font(name="Courier New", bold=True)
    ws["D7"].font = Font(name="Courier New", bold=True)
    ws["D7"].alignment = Alignment(horizontal="right")

    headers = ["SR. NO", "DESCRIPTION", "QTY", "RATE (₹)"]
    header_row = 9
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Courier New", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

    row_num = 10
    for idx, item in enumerate(items, 1):
        sr = item.get("sr_no", idx)
        desc = item.get("description", "")
        qty = item.get("quantity", 1)
        unit = item.get("unit", "NO")
        rate_val = rates.get(str(sr)) or rates.get(sr) or 0

        ws.cell(row=row_num, column=1, value=sr).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=3, value=f"{qty} {unit}").alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=float(rate_val) if rate_val else 0).alignment = Alignment(horizontal="right")

        for c in range(1, 5):
            ws.cell(row=row_num, column=c).border = thick_border
            ws.cell(row=row_num, column=c).font = Font(name="Courier New", size=9, bold=True)
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="(1) GST@18% Extra\n(2) For Destination\n(3) Delivery within 30 days\n(4) Inspection by consignees\n(5) Payment 100% against CRN").font = Font(name="Courier New", size=8, bold=True)
    ws.cell(row=row_num, column=4, value="M/S RAJU ENGINEERING WORKS\n\nProprietor").font = Font(name="Courier New", bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="right")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    return output_path

