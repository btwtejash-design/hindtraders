import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, Any

def generate_challan_excel(data: Dict[str, Any], output_path: str) -> str:
    """
    Generates Delivery Challan Excel matching sample Challan.xlsx.
    Uses dynamic Consignee from PO for the "To" block.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = str(data.get("challan_no", "37"))

    ws.views.sheetView[0].showGridLines = True

    bold_font = Font(name="Arial", size=10, bold=True)
    header_font = Font(name="Arial", size=14, bold=True)
    regular_font = Font(name="Arial", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    # Top Row headers
    ws['A1'] = f"GSTIN:- {data['vendor']['gstin']}"
    ws['A1'].font = bold_font

    ws.merge_cells('B1:C1')
    ws['B1'] = "CHALLAN"
    ws['B1'].font = header_font
    ws['B1'].alignment = center_align

    ws['D1'] = f"Mob:- {data['vendor']['phone']}"
    ws['D1'].font = bold_font
    ws['D1'].alignment = right_align

    # Company Header
    ws.merge_cells('A3:D3')
    ws['A3'] = data['vendor']['name'].upper()
    ws['A3'].font = Font(name="Arial", size=16, bold=True)
    ws['A3'].alignment = center_align

    ws.merge_cells('A4:D4')
    ws['A4'] = "RAILWAY CONTRACTOR & SUPPLIER"
    ws['A4'].font = bold_font
    ws['A4'].alignment = center_align

    ws.merge_cells('A5:D5')
    ws['A5'] = "Manufactures:-  Diesel Locomotives Spare Parts, Ferrous & Non Ferrous Components and General Order Suppliers"
    ws['A5'].font = Font(name="Arial", size=8, italic=True)
    ws['A5'].alignment = center_align

    ws.merge_cells('A6:D6')
    ws['A6'] = data['vendor']['address'].upper()
    ws['A6'].font = bold_font
    ws['A6'].alignment = center_align

    # Challan No & Date
    ws['B10'] = f"Challan No:- {data.get('invoice_no', data.get('challan_no', '42'))}"
    ws['B10'].font = bold_font

    ws['D10'] = f"Date:- {data.get('challan_date', '11/04/2026')}"
    ws['D10'].font = bold_font
    ws['D10'].alignment = right_align

    # Dynamic "To" Section (Extracted Consignee from PO)
    consignee_str = data.get("consignee", "SSE / DPS")
    to_block = (
        f"To,\n"
        f"     {consignee_str}\n"
        f"     Eastern Rly. Jamalpur\n"
        f"     JAMALPUR, 811214\n"
    )
    ws.merge_cells('B12:C15')
    ws['B12'] = to_block
    ws['B12'].font = bold_font
    ws['B12'].alignment = left_align

    # PO No & Date
    ws['B17'] = f"Purchase order No. {data.get('po_number', '55265300100493')}"
    ws['B17'].font = bold_font

    ws['D17'] = f"Date:- {data.get('po_date', '14/03/2026')}"
    ws['D17'].font = bold_font
    ws['D17'].alignment = right_align

    # Items Table Header Row 20
    ws['B20'] = "Sr. No"
    ws['B20'].font = bold_font
    ws['B20'].alignment = center_align
    ws['B20'].border = thin_border
    ws['B20'].fill = header_fill

    ws['C20'] = "Particulars"
    ws['C20'].font = bold_font
    ws['C20'].alignment = center_align
    ws['C20'].border = thin_border
    ws['C20'].fill = header_fill

    ws['D20'] = "Quantity"
    ws['D20'].font = bold_font
    ws['D20'].alignment = center_align
    ws['D20'].border = thin_border
    ws['D20'].fill = header_fill

    current_row = 21
    for item in data.get("items", []):
        ws.cell(row=current_row, column=2, value=item["sr_no"]).alignment = center_align
        ws.cell(row=current_row, column=2).font = regular_font
        ws.cell(row=current_row, column=2).border = thin_border

        desc_text = item["description"]
        ws.cell(row=current_row, column=3, value=desc_text).alignment = left_align
        ws.cell(row=current_row, column=3).font = regular_font
        ws.cell(row=current_row, column=3).border = thin_border

        qty_str = item.get("quantity_display", f"{item['quantity']} Nos")
        ws.cell(row=current_row, column=4, value=qty_str).alignment = center_align
        ws.cell(row=current_row, column=4).font = regular_font
        ws.cell(row=current_row, column=4).border = thin_border

        current_row += 3

    # Footer
    footer_row = max(current_row + 4, 38)
    ws[f'A{footer_row}'] = "Receiver Signature"
    ws[f'A{footer_row}'].font = bold_font

    ws[f'A{footer_row+3}'] = "1. All legal proceeding by or against us shall be instituted in Munger County only."
    ws[f'A{footer_row+3}'].font = Font(name="Arial", size=8)

    ws[f'A{footer_row+4}'] = "2. Goods once sold can not be returned or Exchange."
    ws[f'A{footer_row+4}'].font = Font(name="Arial", size=8)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
