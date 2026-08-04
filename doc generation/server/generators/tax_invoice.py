import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, Any
from server.generators.num_to_words import amount_to_words

def generate_tax_invoice_excel(data: Dict[str, Any], output_path: str) -> str:
    """
    Generates Tax Invoice Excel file matching Tax Invoice(Empty Format)1.xlsx structure.
    Details box width is compact and aligned with right columns.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"

    ws.views.sheetView[0].showGridLines = True

    bold_font = Font(name="Arial", size=9.5, bold=True)
    header_font = Font(name="Arial", size=14, bold=True)
    regular_font = Font(name="Arial", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Set column widths - kept compact for Details block
    col_widths = {'A': 6, 'B': 45, 'C': 8, 'D': 10, 'E': 11, 'F': 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 1: Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "TAX INVOICE"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align

    # Details Block (Top Right - Narrow & Compact)
    ws.merge_cells('D3:F3')
    ws['D3'] = "DETAILS"
    ws['D3'].font = bold_font
    ws['D3'].alignment = center_align
    ws['D3'].fill = header_fill
    ws['D3'].border = thin_border

    details = [
        ("INVOICE NO:", data.get("invoice_no", "42")),
        ("DATE:", data.get("invoice_date", "15-06-2026")),
        ("PO.NO:", data.get("po_number", "55265692101304")),
        ("DATE:", data.get("po_date", "13-07-2026")),
        ("DELIVERY CHAALAN NO:", data.get("challan_no", "44")),
        ("DATE:", data.get("challan_date", "24-07-2026")),
        ("CRN.No- " + str(data.get('crn_no', '') or ''), "DATE: " + str(data.get('crn_date', '') or ''))
    ]

    for idx, (k, v) in enumerate(details, start=4):
        if idx == 10:  # CRN row
            ws['D10'] = k
            ws['D10'].font = bold_font
            ws['D10'].border = thin_border
            ws.merge_cells('E10:F10')
            ws['E10'] = v
            ws['E10'].font = bold_font
            ws['E10'].border = thin_border
        else:
            ws.merge_cells(f'D{idx}:E{idx}')
            ws[f'D{idx}'] = k
            ws[f'D{idx}'].font = bold_font
            ws[f'D{idx}'].border = thin_border
            ws[f'F{idx}'] = str(v)
            ws[f'F{idx}'].font = regular_font
            ws[f'F{idx}'].border = thin_border

    # FROM Block
    ws.merge_cells('A12:C12')
    ws['A12'] = "FROM"
    ws['A12'].font = bold_font
    ws['A12'].fill = header_fill
    ws['A12'].border = thin_border

    from_text = (
        f"COMPANY: {data['vendor']['name']}\n"
        f"ADDRESS: {data['vendor']['address']}\n"
        f"PHONE: {data['vendor']['phone']}\n"
        f"E-MAIL: {data['vendor']['email']}\n"
        f"GSTIN: {data['vendor']['gstin']}"
    )
    ws.merge_cells('A13:C18')
    ws['A13'] = from_text
    ws['A13'].alignment = left_align
    ws['A13'].font = regular_font
    ws['A13'].border = thin_border

    # BILL TO Block
    ws.merge_cells('D12:F12')
    ws['D12'] = "BILL TO"
    ws['D12'].font = bold_font
    ws['D12'].fill = header_fill
    ws['D12'].border = thin_border

    bill_to_text = (
        f"TO: {data['bill_to']['name']}\n"
        f"DEPARTMENT: {data['bill_to']['department']}\n"
        f"LOCATION: {data['bill_to']['location']}\n"
        f"CONSIGNEE: {data['bill_to']['consignee']}\n"
        f"STATE CODE: {data['bill_to']['state_code']}"
    )
    ws.merge_cells('D13:F18')
    ws['D13'] = bill_to_text
    ws['D13'].alignment = left_align
    ws['D13'].font = regular_font
    ws['D13'].border = thin_border

    # Table Header Row 20
    headers = ["Sr. No.", "Description", "HSN", "Quantity (Nos)", "Rate/ Unit (₹)", "Total (₹)"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=20, column=col_idx, value=text)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    current_row = 21
    taxable_sum = 0.0

    for item in data.get("items", []):
        ws.cell(row=current_row, column=1, value=item["sr_no"]).alignment = center_align
        ws.cell(row=current_row, column=2, value=item["description"]).alignment = left_align
        ws.cell(row=current_row, column=3, value=item["hsn"]).alignment = center_align
        ws.cell(row=current_row, column=4, value=item["quantity"]).alignment = center_align
        ws.cell(row=current_row, column=5, value=item["rate"]).alignment = right_align
        
        tot = round(float(item["quantity"]) * float(item["rate"]), 2)
        taxable_sum += tot
        ws.cell(row=current_row, column=6, value=tot).alignment = right_align

        for c in range(1, 7):
            ws.cell(row=current_row, column=c).border = thin_border
            ws.cell(row=current_row, column=c).font = regular_font

        current_row += 3

    # Summary Row
    gst_rate = data.get("items", [{}])[0].get("gst_percent", 18.0)
    cgst_rate = gst_rate / 2.0
    sgst_rate = gst_rate / 2.0

    cgst_amount = round(taxable_sum * (cgst_rate / 100.0), 2)
    sgst_amount = round(taxable_sum * (sgst_rate / 100.0), 2)
    grand_total = round(taxable_sum + cgst_amount + sgst_amount, 2)
    words_total = amount_to_words(grand_total)

    summary_row = max(current_row + 2, 35)

    ws.merge_cells(f'A{summary_row}:C{summary_row+3}')
    ws[f'A{summary_row}'] = f"Grand Total in Words:  {words_total}"
    ws[f'A{summary_row}'].font = bold_font
    ws[f'A{summary_row}'].alignment = left_align
    ws[f'A{summary_row}'].border = thin_border

    labels = [
        ("Taxable Amount:", taxable_sum),
        (f"SGST: {int(sgst_rate)}%", sgst_amount),
        (f"CGST: {int(cgst_rate)}%", cgst_amount),
        ("Grand Total:", grand_total)
    ]

    for idx, (label, val) in enumerate(labels):
        r = summary_row + idx
        ws.merge_cells(f'D{r}:E{r}')
        ws[f'D{r}'] = label
        ws[f'D{r}'].font = bold_font
        ws[f'D{r}'].border = thin_border
        
        ws[f'F{r}'] = val
        ws[f'F{r}'].font = bold_font
        ws[f'F{r}'].alignment = right_align
        ws[f'F{r}'].border = thin_border

    acc_row = summary_row + 6
    ws[f'A{acc_row}'] = "Account Details-"
    ws[f'A{acc_row}'].font = bold_font

    bank_details = [
        ("Bank:", data['vendor']['bank_name']),
        ("Account No:", data['vendor']['account_no']),
        ("IFSC Code:", data['vendor']['ifsc']),
        ("Branch:", data['vendor']['branch'])
    ]

    for idx, (k, v) in enumerate(bank_details, start=acc_row+1):
        ws[f'A{idx}'] = k
        ws[f'A{idx}'].font = bold_font
        ws[f'C{idx}'] = v
        ws[f'C{idx}'].font = regular_font

    ws[f'F{acc_row+4}'] = "-Proprietor"
    ws[f'F{acc_row+4}'].font = bold_font
    ws[f'F{acc_row+4}'].alignment = right_align

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
